#!/usr/bin/env python3

import argparse
import crossbow as cxb
import pyarrow as pa
import pyarrow.csv as csv
import pandas as pd
from openpyxl import load_workbook
from typing import Optional
import os
import sys

venv_path = "/path/to/your/venv/bin/activate_this.py"
with open(venv_path) as f:
    exec(f.read(), {"__file__": venv_path})


def select_excel_file() -> Optional[str]:

    xlsx_files = [f for f in os.listdir(os.getcwd()) if f.endswith(".xlsx")]

    if not xlsx_files:
        print("No Excel files found in the current directory.")
        return None

    print("Select an Excel file to read:")
    for i, file_name in enumerate(xlsx_files, 1):
        print(f"{i}. {file_name}")

    while True:
        try:
            choice = int(input("Enter the number corresponding to the file: "))
            if 1 <= choice <= len(xlsx_files):
                return xlsx_files[choice - 1]
            else:
                print(f"Please enter a number between 1 and {len(xlsx_files)}.")
        except ValueError:
            print("Invalid input. Please enter a valid number.")


def get_sheet_names(file_path: str) -> list[str]:

    workbook = load_workbook(filename=file_path, read_only=True)

    sheet_names = []
    for sheet in workbook.sheetnames:
        sheet_names.append(sheet)

    return sheet_names
